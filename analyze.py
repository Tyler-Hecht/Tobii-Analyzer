import glob
import os
import glob
import pandas as pd
import openpyxl as op
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from copy import copy
import subprocess
import shutil

try:
    os.mkdir("OUTPUT")
except:
    pass

if os.name == "nt":
    sep = "\\"
else:
    sep = "/"
path = os.getcwd()
input_path = path + "/INPUT/"
output_path = path + "/OUTPUT/"

# get input
input_files = glob.glob(os.path.join(input_path, "*.tsv"))
if len(input_files) > 1:
    print("ERROR: Multiple inputs files found")
    exit(1)
elif not input_files:
    print("ERROR: No .tsv files found in INPUT folder")
    exit(1)
else:
    input_file = input_files[0].split(sep)[-1]
input_as_csv = input_file.split(".")[0] + ".csv"

with open('config.txt') as f:
    lines = f.readlines()
try:
    order = lines[1].strip().upper()
except:
    print("ERROR: No order given in config file")
    exit(1)
try:
    t = lines[3].strip().lower()
except:
    print("ERROR: No testing or training given in config file")
    exit(1)
try:
    ms = lines[5].strip()
except:
    print("ERROR: No delay given in config file")
    exit(1)
    
# run the Java file
os.chdir(input_path)
subprocess.Popen(["javac", "Tobii.java"])
p = subprocess.Popen(["java", "Tobii", t, input_file, input_as_csv, ms])
p.communicate()

input_files = glob.glob(os.path.join(input_path, "*.csv"))
input_file = input_files[0]

input_data = pd.read_csv(input_file, keep_default_na=False)
os.chdir(path)
m_orders_data = pd.read_excel("orders.xlsx", header = 1, usecols = [1, 3, 5, 7], sheet_name = 2, nrows = 21)
b_orders_data = pd.read_excel("orders.xlsx", header = 26, usecols = [1, 3, 5, 7], sheet_name = 2, nrows = 21)
orders_data = m_orders_data.join(b_orders_data)

try:
    order_data = orders_data[order]
except:
    print("ERROR: Invalid order " + order + " given in config file")
    exit(1)

# create dictionary of information
info = {}
for trial in order_data:
    info[trial] = {}
    for i in range(0, len(input_data["Trial"])):
        aoi_hit = input_data["Trial"][i]
        props = input_data["Proportion"]
        frames = input_data["Frames"]
        times = input_data["Time (ms)"]
        if not aoi_hit:
            continue
        if trial in aoi_hit:
            if "target" in aoi_hit:
                if trial.split("_")[-1] == "R":
                    info[trial]["Right prop"] = round(props[i], 6)
                    info[trial]["Right Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
                else:
                    info[trial]["Left prop"] = round(props[i], 6)
                    info[trial]["Left Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
            elif "distractor" in aoi_hit:
                if trial.split("_")[-1] == "R":
                    info[trial]["Left prop"] = round(props[i], 6)
                    info[trial]["Left Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
                else:
                    info[trial]["Right prop"] = round(props[i], 6)
                    info[trial]["Right Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
            elif "BASELINE" in aoi_hit:
                aoi_hit_sides = aoi_hit.split("[")[1].split("_")[0].split("-")
                sides = {aoi_hit_sides[0]: "left", aoi_hit_sides[1]: "right"}
                additional = aoi_hit.split()[-1][:-1]
                if sides[additional] == "left":
                    info[trial]["Left prop"] = round(props[i], 6)
                    info[trial]["Left Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
                else:
                    info[trial]["Right prop"] = round(props[i], 6)
                    info[trial]["Right Frames/Time"] = str(frames[i]) + " / " + str(round(times[i], 2))
# make information into data frames
left_ft = []
right_ft = []
left_props = []
right_props = []
for trial in order_data:
    if trial == "Filler-Video":
        left_ft.append("")
        right_ft.append("")
        left_props.append("")
        right_props.append("")
    else:
        left_ft.append((info[trial]["Left Frames/Time"]))
        right_ft.append((info[trial]["Right Frames/Time"]))
        left_props.append((info[trial]["Left prop"]))
        right_props.append((info[trial]["Right prop"]))
lrdf = pd.DataFrame(data = {"Left Frames/Time": left_ft, "Right Frames/Time": right_ft, "Left Prop": left_props, "Right Prop": right_props})
output_data = lrdf.join(order_data)

# write to Excel
name = input_file.split(sep)[-1].split(".")[0]
new_name = name+"_organized.xlsx"
os.chdir(output_path)
output_data.to_excel(new_name, sheet_name=name, index = False)

# colorize
output_wb = op.load_workbook(new_name)
output_sheet = list(output_wb)[0]
os.chdir(path)
orders_wb = op.load_workbook("orders.xlsx")
orders_sheet = list(orders_wb)[2]
if "M" in order:
    start = 2
else:
    start = 27
if "1" in order:
    col = "B"
elif "2" in order:
    col = "D"
elif "3" in order:
    col = "F"
else:
    col = "H"
for i in range(1, len(output_sheet["E"])):
    output_cell_num = "E" + str(i + 1)
    orders_cell_num = col + str(start + i)
    output_sheet[output_cell_num].fill = copy(orders_sheet[orders_cell_num].fill)
os.chdir(output_path)
# make it look nicer
output_sheet["A1"].font = Font(bold=False)
output_sheet["B1"].font = Font(bold=False)
output_sheet["C1"].font = Font(bold=False)
output_sheet["D1"].font = Font(bold=False)
output_sheet["E1"].font = Font(bold=False)
output_sheet["E1"].value = "Trial"
output_sheet.column_dimensions["A"].width = 20
output_sheet.column_dimensions["B"].width = 20
output_sheet.column_dimensions["C"].width = 15
output_sheet.column_dimensions["D"].width = 15
output_sheet.column_dimensions["E"].width = 30
for row in output_sheet:
    for cell in row:
        cell.font = Font(size = "12")
        cell.alignment = Alignment(horizontal='right')
for cell in output_sheet["E"]:
    cell.alignment = Alignment(horizontal='left')
    cell.border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
output_wb.save(new_name)

os.chdir(input_path)
input_as_csv2 = input_as_csv.split(".")[0] + "_analyzed.csv"
shutil.move(input_as_csv, output_path + input_as_csv2) 
